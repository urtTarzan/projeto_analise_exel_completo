# ─── DOCUMENTAÇÃO E COMENTÁRIOS ─────────────────────────────────────────
# Sistema de Processamento de Arquivos com Validação, Filtragem e Exportação
# Autor: Matheus Silva Sousa oliveira / @urtTarzan GITHUB
# Objetivo: Automatizar o processamento de arquivos (xlsx, csv, json), validar colunas,
# mover arquivos com erro, destacar CPFs inválidos e gerar relatórios organizados.

from openpyxl import load_workbook
from openpyxl.styles import PatternFill
from openpyxl.utils import get_column_letter
from shutil import move
import pandas as pd
import logging
import os

# ───── Criação de Pastas Necessárias ─────────────────────────
# A pasta de relatórios é criada primeiro para garantir o funcionamento do logger
if not os.path.exists("relatorios"):
    os.mkdir("relatorios")

# ───── Configuração do Logger ─────────────────────────
logger = logging.getLogger()
logger.setLevel(logging.INFO)

formatter = logging.Formatter('%(asctime)s - %(levelname)s - %(message)s')
file_handler = logging.FileHandler('relatorios/processamento.log', mode='a', encoding='utf-8')
file_handler.setFormatter(formatter)
console_handler = logging.StreamHandler()
console_handler.setFormatter(formatter)
logger.addHandler(file_handler)
logger.addHandler(console_handler)

# ───── Criação das Pastas Base ─────────────────────────
interromper_caso_não_existir = False

if not os.path.exists("arquivos"):
    os.mkdir("arquivos")
    interromper_caso_não_existir = True

if not os.path.exists("arquivos/brutos"):
    os.mkdir("arquivos/brutos")
    interromper_caso_não_existir = True

if not os.path.exists("arquivos/organizados"):
    os.mkdir("arquivos/organizados")


if interromper_caso_não_existir:
    logger.error("ERRO_FALTA: devido a falta das pastas principais o programa foi parado, rode novamente para funcionar ")
    exit()

# ───── Verificação se há arquivos na pasta ─────────────────────────
pasta_brutos = "arquivos/brutos"
tem_arquivos = any(os.scandir(pasta_brutos))
if not tem_arquivos:
    logging.error("A pasta 'arquivos/brutos' está vazia. Nenhum arquivo será processado.")
# ───── Função para Carregar Arquivos ─────────────────────────
def carregar_arquivo(caminho):
    extensao = caminho.split('.')[-1]
    try:
        if extensao == "json":
            return pd.read_json(caminho)
        elif extensao == "xlsx":
            return pd.read_excel(caminho, engine="openpyxl")
        elif extensao == "csv":
            return pd.read_csv(caminho)
        else:
            logging.warning(f"Extensão não suportada: {caminho}")
            return None
    except Exception as e:
        logging.error(f"Erro ao carregar {caminho}: {e}")
        return None

# ───── Validação de Colunas ─────────────────────────
def validar_colunas(df):
    colunas_esperadas = {'Nome', 'CPF', 'Data', 'Valor', 'Status', 'Tipo de Contrato'}
    return colunas_esperadas.issubset(df.columns)

# ───── Destacar CPFs Inválidos ─────────────────────────
def destacar_cpfs_invalidos(path):
    wb = load_workbook(path)
    ws = wb.active
    coluna_cpf = None

    for idx, cell in enumerate(ws[1], start=1):
        if cell.value == 'CPF':
            coluna_cpf = idx
            break

    if not coluna_cpf:
        logging.warning("Coluna CPF não encontrada.")
        return 0

    vermelho = PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid')
    contador = 0

    for row in ws.iter_rows(min_row=2, min_col=coluna_cpf, max_col=coluna_cpf):
        for cell in row:
            cpf = str(cell.value)
            cpf_numerico = ''.join(filter(str.isdigit, cpf))
            if len(cpf_numerico) != 11:
                cell.fill = vermelho
                contador += 1

    wb.save(path)
    return contador

# ───── Ajustar Largura das Colunas ─────────────────────────
def ajustar_largura_colunas(path):
    wb = load_workbook(path)
    ws = wb.active

    for col in ws.columns:
        max_length = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            if cell.value:
                valor = str(cell.value)
                max_length = max(max_length, len(valor))
        ws.column_dimensions[col_letter].width = max_length + 2

    wb.save(path)

# ───── Função Principal ─────────────────────────
def codigo_principal():
    lista_analisados = []
    for raiz, pastas, arquivos in os.walk("arquivos/organizados"):
        for arquivo in arquivos:
            lista_analisados.append(arquivo.rsplit('.', 1)[0])

    for raiz, pastas, arquivos in os.walk("arquivos/brutos"):
        for arquivo in arquivos:

            if arquivo.rsplit('.', 1)[0] not in lista_analisados:
                logging.info(f"///"*20)
                logging.info(f"Arquivo atual: {arquivo}")
                caminho = os.path.join("arquivos", "brutos", arquivo)

                df = carregar_arquivo(caminho)
                if df is None:
                    continue

                if not validar_colunas(df):
                    logging.warning(f"Colunas faltando no arquivo {arquivo}.")
                    continue

                # Limpeza de dados
                df = df.drop_duplicates()
                df['Nome'] = df['Nome'].str.title()
                #textando a validação do cpf pois nele da pra verificar se o arquivo ja foi organizado
                #// talvez não sera nessesario no codigo final VERIFICAR
                try:
                    df['CPF'] = df['CPF'].str.replace(r'\D', '', regex=True)
                except AttributeError as erro:
                    logging.warning(f"O ARQUIVO {arquivo} JA FOI ANALISADO")
                    continue
                df['Data'] = pd.to_datetime(df['Data'])

                # Filtro de clientes ativos ou cancelados
                if "Inadimplente" in df['Status'].values:
                    df_ativos = df[(df['Status'] == 'Inadimplente') & (df['Valor'] > 1000)]
                else:
                    ano_referencia = 2025
                    df_ativos = df[(df['Status'] == 'Cancelado') & (df['Valor'] > 4000) & (df["Data"].dt.year == ano_referencia)]

                # Agrupamento por tipo de contrato
                resumo_contratos = df_ativos.groupby('Tipo de Contrato').agg({
                    'Valor': ['count', 'sum']
                }).reset_index()
                resumo_contratos.columns = ['Tipo de Contrato', 'Qtd Clientes', 'Valor Total (R$)']

                # Exportações
                if not arquivo.endswith("xlsx"):
                    arquivo = arquivo.rsplit(".", 1)[0] + ".xlsx"

                resumo_path = os.path.join("relatorios", arquivo)
                organizados_path = os.path.join("arquivos", "organizados", arquivo)

                df_ativos.to_excel(organizados_path, index=False)
                resumo_contratos.to_excel(resumo_path, index=False)

                logger.info(f"{arquivo} exportado com {len(df_ativos)} clientes ativos.")

                # CPFs inválidos e largura
                invalidos = destacar_cpfs_invalidos(organizados_path)
                logger.info(f"{invalidos} CPFs inválidos foram destacados.")
                ajustar_largura_colunas(organizados_path)