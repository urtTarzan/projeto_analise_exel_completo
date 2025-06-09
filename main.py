import pandas as pd
import os
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
from openpyxl.utils import get_column_letter
import logging

# pasta relatorio sendo criada antes pois sem ela não tem logger.

if not os.path.exists("relatorios"):
    os.mkdir("relatorios")

# ───── Logger Config ─────────────────────────
logger = logging.getLogger()
logger.setLevel(logging.INFO)

formatter = logging.Formatter('%(asctime)s - %(levelname)s - %(message)s')

file_handler = logging.FileHandler('relatorios/processamento.log', mode='a', encoding='utf-8')
file_handler.setFormatter(formatter)

console_handler = logging.StreamHandler()
console_handler.setFormatter(formatter)

logger.addHandler(file_handler)
logger.addHandler(console_handler)



# ____ crinado pasta se não existir______________

interromper_caso_não_existir = False


if not os.path.exists("arquivos"):
    os.mkdir("arquivos")
    interromper_caso_não_existir = True

if not os.path.exists("arquivos/brutos"):
    os.mkdir("arquivos/brutos")
    interromper_caso_não_existir = True

if not os.path.exists("arquivos/organizados"):
    os.mkdir("arquivos/organizados")
    
if  interromper_caso_não_existir == True:
    logger.error("ERRO_FALTA: devido a falta das pastas principais o programa foi parado ")

#testando se os arquivos estão vazios
pasta_brutos = "arquivos/brutos"
tem_arquivos = any(os.scandir(pasta_brutos))

if not tem_arquivos:
    logging.warning("A pasta 'arquivos/brutos' está vazia. Nenhum arquivo será processado.")
    exit()
    

# ───── Funções Utilitárias ─────────────────────────

def carregar_arquivo(caminho):
    extensao = caminho.split('.')[-1]
    try:
        if extensao == "json":
            return pd.read_json(caminho)
        elif extensao == "xlsx":
            return pd.read_excel(caminho, engine="openpyxl")
        elif extensao == "csv":
            return pd.read_csv(caminho)
        else:
            logging.error(f"Extensão não suportada: {caminho}")
            return None
    except Exception as e:
        logging.error(f"Erro ao carregar {caminho}: {e}")
        return None

def validar_colunas(df):
    colunas_esperadas = {'Nome', 'CPF', 'Data', 'Valor', 'Status', 'Tipo de Contrato'}
    return colunas_esperadas.issubset(df.columns)

def destacar_cpfs_invalidos(path):
    wb = load_workbook(path)
    ws = wb.active
    coluna_cpf = None

    for idx, cell in enumerate(ws[1], start=1):
        if cell.value == 'CPF':
            coluna_cpf = idx
            break

    if not coluna_cpf:
        logging.warning("Coluna CPF não encontrada.")
        return 0

    vermelho = PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid')
    contador = 0

    for row in ws.iter_rows(min_row=2, min_col=coluna_cpf, max_col=coluna_cpf):
        for cell in row:
            cpf = str(cell.value)
            cpf_numerico = ''.join(filter(str.isdigit, cpf))
            if len(cpf_numerico) != 11:
                cell.fill = vermelho
                contador += 1

    wb.save(path)
    return contador

def ajustar_largura_colunas(path):
    wb = load_workbook(path)
    ws = wb.active

    for col in ws.columns:
        max_length = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            if cell.value:
                valor = str(cell.value)
                max_length = max(max_length, len(valor))
        ws.column_dimensions[col_letter].width = max_length + 2

    wb.save(path)


# ───── Processo Principal ─────────────────────────

for raiz, pastas, arquivos in os.walk("arquivos/brutos"):

    for arquivo in arquivos:
        logging.info(f"Arquivo atual: {arquivo}")
        caminho = os.path.join("arquivos", "brutos", arquivo)

        df = carregar_arquivo(caminho)
        if df is None:
            continue

        if not validar_colunas(df):
            logging.error(f"Colunas faltando no arquivo {arquivo}. Pulando.")
            continue

        # Limpeza de dados
        df = df.drop_duplicates()
        df['Nome'] = df['Nome'].str.title()
        df['CPF'] = df['CPF'].str.replace(r'\D', '', regex=True)
        df['Data'] = pd.to_datetime(df['Data'])

        # Filtro de clientes ativos ou cancelados
        if "Inadimplente" in df['Status'].values:
            df_ativos = df[(df['Status'] == 'Inadimplente') & (df['Valor'] > 1000)]
        else:
            ano_referencia = 2025
            df_ativos = df[
            (df['Status'] == 'Cancelado') &
            (df['Valor'] > 4000) &
            (df["Data"].dt.year == ano_referencia)
        ]
                # Agrupamento por tipo de contrato
        resumo_contratos = df_ativos.groupby('Tipo de Contrato').agg({
            'Valor': ['count', 'sum']
        }).reset_index()
        resumo_contratos.columns = ['Tipo de Contrato', 'Qtd Clientes', 'Valor Total (R$)']

        # Preparar paths de exportação
        if not arquivo.endswith("xlsx"):
            arquivo = arquivo.rsplit(".", 1)[0] + ".xlsx"
        resumo_path = os.path.join("relatorios", arquivo)
        organizados_path = os.path.join("arquivos", "organizados", arquivo)

        # Exportações
        df_ativos.to_excel(organizados_path, index=False)
        resumo_contratos.to_excel(resumo_path, index=False)

        logger.info(f"{arquivo} exportado com {len(df_ativos)} clientes ativos.")

        # Destacar CPFs inválidos
        invalidos = destacar_cpfs_invalidos(organizados_path)
        logger.warning(f"{invalidos} CPFs inválidos foram destacados.")

        # Ajustar largura das colunas
        ajustar_largura_colunas(organizados_path)
